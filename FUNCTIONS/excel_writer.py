from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from typing import Dict, List, Optional, Union
import pandas as pd
import os

class ExcelWriter:
    def __init__(self, file_path: str):
        """
        Initialize ExcelWriter with the path to the Excel file.
        
        Args:
            file_path (str): Path to the Excel file (.xlsx)
        """
        self.file_path = file_path
        self._validate_file()
        self.workbook = None
        self.sheet = None
        self.header_row = 1  # Assuming headers are in first row
        self._load_workbook()

    def _validate_file(self) -> None:
        """Validate if the file exists and has the correct extension."""
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f"File not found: {self.file_path}")
        
        if not self.file_path.endswith('.xlsx'):
            raise ValueError("File must be .xlsx format")

    def _load_workbook(self) -> None:
        """Load the workbook and get the active sheet."""
        try:
            self.workbook = load_workbook(self.file_path)
            self.sheet = self.workbook.active
        except Exception as e:
            raise Exception(f"Error loading workbook: {str(e)}")

    def _get_column_index(self, column_name: str) -> int:
        """Get the column index for a given column name."""
        for idx, cell in enumerate(self.sheet[self.header_row], 1):
            if cell.value == column_name:
                return idx
        raise ValueError(f"Column '{column_name}' not found")

    def _get_row_index(self, search_column: str, search_value: str) -> Optional[int]:
        """
        Find the row index for a given value in a specific column.
        Returns None if not found.
        """
        col_idx = self._get_column_index(search_column)
        for row_idx in range(self.header_row + 1, self.sheet.max_row + 1):
            if str(self.sheet.cell(row=row_idx, column=col_idx).value) == str(search_value):
                return row_idx
        return None

    def append_row(self, data: Dict[str, Union[str, int, float]]) -> None:
        """
        Append a new row to the Excel file.
        
        Args:
            data (Dict): Dictionary containing column names and values
        """
        try:
            # Validate all columns exist
            for col in data.keys():
                self._get_column_index(col)

            # Get next row index
            next_row = self.sheet.max_row + 1

            # Write data
            for col_name, value in data.items():
                col_idx = self._get_column_index(col_name)
                self.sheet.cell(row=next_row, column=col_idx, value=value)

            self.workbook.save(self.file_path)
        except Exception as e:
            raise Exception(f"Error appending row: {str(e)}")

    def update_row(self, search_column: str, search_value: str, 
                  new_data: Dict[str, Union[str, int, float]]) -> bool:
        """
        Update an existing row in the Excel file.
        
        Args:
            search_column (str): Column name to search in
            search_value (str): Value to search for
            new_data (Dict): Dictionary containing column names and new values
            
        Returns:
            bool: True if row was updated, False if not found
        """
        try:
            row_idx = self._get_row_index(search_column, search_value)
            if row_idx is None:
                return False

            # Validate all columns exist
            for col in new_data.keys():
                self._get_column_index(col)

            # Update data
            for col_name, value in new_data.items():
                col_idx = self._get_column_index(col_name)
                self.sheet.cell(row=row_idx, column=col_idx, value=value)

            self.workbook.save(self.file_path)
            return True
        except Exception as e:
            raise Exception(f"Error updating row: {str(e)}")

    def delete_row(self, search_column: str, search_value: str) -> bool:
        """
        Delete a row from the Excel file.
        
        Args:
            search_column (str): Column name to search in
            search_value (str): Value to search for
            
        Returns:
            bool: True if row was deleted, False if not found
        """
        try:
            row_idx = self._get_row_index(search_column, search_value)
            if row_idx is None:
                return False

            # Delete the row
            self.sheet.delete_rows(row_idx)
            self.workbook.save(self.file_path)
            return True
        except Exception as e:
            raise Exception(f"Error deleting row: {str(e)}")

    def get_row(self, search_column: str, search_value: str) -> Optional[Dict]:
        """
        Get a row from the Excel file.
        
        Args:
            search_column (str): Column name to search in
            search_value (str): Value to search for
            
        Returns:
            Optional[Dict]: Dictionary containing row data if found, None otherwise
        """
        try:
            row_idx = self._get_row_index(search_column, search_value)
            if row_idx is None:
                return None

            # Get column headers
            headers = [cell.value for cell in self.sheet[self.header_row]]
            
            # Get row data
            row_data = {}
            for col_idx, header in enumerate(headers, 1):
                value = self.sheet.cell(row=row_idx, column=col_idx).value
                row_data[header] = value

            return row_data
        except Exception as e:
            raise Exception(f"Error getting row: {str(e)}")

    def close(self) -> None:
        """Close the workbook."""
        if self.workbook:
            self.workbook.close() 